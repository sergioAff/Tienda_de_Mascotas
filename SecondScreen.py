from tkinter import *
from tkinter import ttk
from tkinter import messagebox,filedialog
import sqlite3 as sql
import traceback
from datetime import *
import openpyxl
import os
import traceback
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from registros import Registro

class Second_Screen:
    WINDOW_WIDTH = 950
    WINDOW_HEIGHT = 600

    def __init__(self, x, y):
        self.setup_ui(x, y)

    def setup_ui(self, x_position,y_position):
        # Configuración de la interfaz gráfica principal
        self.root = Tk()
        self.root.geometry(f'{self.WINDOW_WIDTH}x{self.WINDOW_HEIGHT}+{x_position}+{y_position}')
        self.root.title('Tienda de Mascotas')
        self.root.resizable(0, 0)
        self.root.config(bg='#082d44')

        
        try:
            # Cargar el archivo de la base de datos
            self.archivo =filedialog.askopenfilename(initialdir=os.getcwd(),title="Seleccione un archivo",filetypes=(("archivos db","*.db"),("all files","*.*")))
            if self.archivo.count('.db')!=1 and self.archivo.count('.sqlite')!=1:
                messagebox.showerror('Error','El archivo seleccionado no es una base de datos valida')
                self.root.destroy()
                return
        except Exception:
            messagebox.showerror('Error','Error al abrir la base de datos')
            self.root.destroy()
            return

        # Carga de las opciones para el menú desplegable
        self.load_options()
        # Creación de los elementos de la interfaz gráfica
        self.create_widgets()
        # Bucle principal de la interfaz gráfica
        self.root.mainloop()

    def create_widgets(self):
        # Creación de los elementos gráficos (botones, frames, etc.)
        self.create_data_frame()
        self.create_tablas_button()

    def create_data_frame(self):
        # Creación del frame para mostrar los datos
        self.frameMostrar = Frame(self.root, bg='#082d44', borderwidth=0, highlightthickness=0)
        self.frameMostrar.place(
            height=500,
            width=940,
            x=20,
            y=130)
        self.frameMostrar.grid_rowconfigure(0, weight=1)
        self.frameMostrar.grid_columnconfigure(0, weight=1)


    def create_tablas_button(self):
        # Creación del botón de tablas
        self.botonTablas = Menubutton(
            text='Tablas',
            cursor='hand2',
            borderwidth=0,
            highlightthickness=0
        )   

        self.botonTablas.menu = Menu(self.botonTablas, tearoff=0)
        self.botonTablas["menu"] = self.botonTablas.menu
        self.botonTablas.menu.config(font=("Helvetica", 14))
        for opcion in self.opciones:
            self.botonTablas.menu.add_command(
                label=opcion,
                command=lambda opt=opcion: self.show_table_and_buttons(opt)
            )
        
        self.botonTablas.place(
            height=23,
            width=100,
            x=20,
            y=25,
        )

    def show_table_and_buttons(self, table_name):
        # Muestra la tabla seleccionada y activa los otros botones
        self.tabla_actual = table_name
        tree = self.tablas(self.frameMostrar, self.archivo, table_name)
        self.create_all_buttons(tree)

        # Verificar si la tabla es de inventario o inventario_animales
        if table_name in ["inventario", "inventario_animales"]:
            # Si ya existe un label del monto total, eliminarlo antes de crear uno nuevo
            if hasattr(self, 'total_label'):
                self.total_label.destroy()

            # Mostrar el monto total
            self.total_label = Label(text=f'Monto Total: {self.calcular_monto_total(table_name)}')
            self.total_label.config(bg='#082d44', font=(['Arial', 14]), fg='white')
            self.total_label.place(
                height=40,
                width=500,
                x=520,
                y=80
            )
        else:
            # Si la tabla no es de inventario, eliminar el label del monto total si existe
            if hasattr(self, 'total_label'):
                self.total_label.destroy()


    def create_all_buttons(self, tree):
        # Creación de los botones de la interfaz gráfica
        self.create_button('Añadir', 32, 110, 150, 22, lambda: self.create_command("Añadir", tree))
        self.create_button('Actualizar', 35, 170, 270, 22, lambda: self.create_command("Actualizar", tree))
        self.create_button('Borrar', 40, 110, 450, 20, lambda: self.create_command("Borrar", tree))
        self.create_button('C.Diario',38,110,600,22,lambda:self.create_command('Guardar',tree))
        self.create_button('C.Semanal',38,130,750,23,lambda:self.create_command('Semanal',tree))


    def create_button(self, text, height, width, x, y, command):
        # Función genérica para crear botones
        button = Label(
            text=text,
            font=("Helvetica", 24),
            cursor='hand2',
            bg='#082d44',
            fg='white',
            # borderwidth=0,
            # highlightthickness=0,
            # command=command
        )
        button.bind("<Button-1>", lambda event: command())

        button.place(
            height=height,
            width=width,
            x=x,
            y=y
        )

    def create_command(self, option, treeview):
        if option == 'Añadir':
            registro = Registro(self.archivo, self.tabla_actual, 'Añadir', self.actualizar_treeview)

        elif option == 'Actualizar':
            val = self.cargar_registro_seleccionado(treeview, 'Actualizar')
            if val is not False:
                registro = Registro(self.archivo, self.tabla_actual, 'Actualizar', self.actualizar_treeview)
                try:
                    registro.cargar(val)
                except:
                    return

        elif option == "Borrar":
            val = self.cargar_registro_seleccionado(treeview, 'Borrar')
            if val is not False:
                self.borrar(val, self.tabla_actual)

        elif option == 'Guardar':
            try:
                
                wb = openpyxl.Workbook()
                ws = wb.active

                # Agregar encabezados de columna
                columnas = [col for col in treeview['columns'] if col not in ['id','id_inventario_animales']]
                    
                ws.append(columnas)

                if self.tabla_actual in ["inventario", "inventario_animales"]:
                    # Agregar monto total como una celda en la primera fila y en la segunda columna del archivo Excel
                    monto_total = self.calcular_monto_total(self.tabla_actual)
                    ws.cell(row=1, column=len(columnas) + 1, value='Importe Total:')
                    ws.cell(row=1, column=len(columnas) + 2, value=monto_total)
                    
                    # Agrupar los registros según la tabla actual
                    resultados_agrupados = self.agrupar_registros(self.tabla_actual)
                    for resultado in resultados_agrupados:
                        ws.append(resultado)                
                    
                else:
                    # Agregar datos de filas
                    for i, item in enumerate(treeview.get_children()):
                        valores_fila = treeview.item(item, 'values')
                        ws.append(valores_fila)

                # Obtener la ruta de la carpeta "registros" en el mismo directorio de la aplicación
                ruta_historial = os.path.join(os.path.dirname(__file__), "historial")

                # Crear la carpeta "historial" si no existe
                if not os.path.exists(ruta_historial):
                    os.makedirs(ruta_historial)
                    
                # Obtener la fecha y hora actual
                ahora = datetime.now()
                hora_actual = ahora.strftime("%H-%M-%S")

                # Guardar el archivo con el nombre de la fecha de hoy
                hoy = datetime.now().strftime("%Y-%m-%d")
                nombre_archivo = f"{self.tabla_actual}_{hoy}_{hora_actual}.xlsx"
                ruta_archivo = os.path.join(ruta_historial, nombre_archivo)
                wb.save(ruta_archivo)
                messagebox.showinfo("Éxito", f"Archivo Excel guardado como {nombre_archivo}")

                if self.tabla_actual in ["inventario", "inventario_animales"]:
                    self.eliminar_todos_los_datos(treeview)

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo generar el archivo Excel: {str(e)}")
        
        elif option == 'Semanal':
            # Verificar si la tabla actual es inventario o inventario_animales
            if self.tabla_actual not in ["inventario", "inventario_animales"]:
                messagebox.showinfo("Alerta", "No se puede hacer cuadre semanal para la tabla actual.")
                return

            # Abrir ventana para seleccionar archivos Excel
            archivos_excel = filedialog.askopenfilenames(
                initialdir=os.getcwd(),
                title="Seleccione uno o más archivos Excel",
                filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
            )

            # Verificar si se seleccionaron archivos
            if not archivos_excel:
                messagebox.showinfo("Alerta", "No se seleccionó ningún archivo Excel.")
                return

            # Variables para almacenar los totales
            total_importe_total = 0
            total_importes_productos = {}
            totales={}

            # Leer cada archivo Excel seleccionado
            for archivo_excel in archivos_excel:
                try:
                    # Cargar el libro de trabajo de Excel
                    wb = load_workbook(archivo_excel)
                    ws = wb.active

                    importe_total = None
                    for row in ws.iter_rows(min_row=1, max_row=1):
                        for cell in row:
                            if cell.value == "Importe Total:":
                                importe_total_cell = ws.cell(row=1, column=cell.column + 1)
                                importe_total = importe_total_cell.value
                                break
                        if importe_total:
                            break
                        
                    if importe_total is None:
                        messagebox.showwarning("Alerta", f"No se encontró el monto total en el archivo Excel {archivo_excel}.")
                        continue  # Pasar al siguiente archivo si no se encontró el monto total
                    
                    # Convertir el monto total a número si es una cadena de texto
                    try:
                        monto_total = float(importe_total)
                    except ValueError:
                        messagebox.showwarning("Alerta", f"El monto total en el archivo Excel {archivo_excel} no es un número válido.")
                        continue  # Pasar al siguiente archivo si el monto total no es un número válido
                    total_importe_total += monto_total

                    # Leer los importes individuales por producto
                    #Aqui va el codigo para leer los importes individuales por producto
                    
                    # Iterar sobre las filas del archivo Excel
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        animal = row[0]
                        producto = row[1]
                        cantidad = row[2]
                        precio = row[3]
                        importe=row[4]
        
                        # Verificar que la cantidad y el importe no sean None
                        if cantidad is not None and importe is not None:
                            # Actualizar los totales para el animal y producto actual
                            clave = (animal, producto, precio)
                            if clave in totales:
                                totales[clave][0] += cantidad
                                totales[clave][1] += importe
                            else:
                                totales[clave] = [cantidad, importe]
                        else:
                            # Si la cantidad o el importe son None, mostrar una advertencia
                            traceback.print_exc()
                            messagebox('Error',"Advertencia: Se encontró un valor None en la cantidad o el importe.")
        
                except Exception as e:
                    traceback.print_exc()
                    messagebox.showerror("Error", f"No se pudo leer el archivo Excel {archivo_excel}: {str(e)}")

            try:
                # Crear la carpeta "cuadre semanal" si no existe
                ruta_cuadre_semanal = os.path.join(os.path.dirname(__file__), "cuadre semanal")
                if not os.path.exists(ruta_cuadre_semanal):
                    os.makedirs(ruta_cuadre_semanal)

                # Obtener la fecha actual
                fecha_actual = datetime.now().strftime("%Y-%m-%d")

                # Crear el nombre del archivo Excel
                nombre_archivo_excel = f"{self.tabla_actual}_{fecha_actual}.xlsx"
                ruta_archivo_excel = os.path.join(ruta_cuadre_semanal, nombre_archivo_excel)

                # Crear un nuevo libro de trabajo de Excel
                wb_nuevo = Workbook()
                ws_nuevo = wb_nuevo.active

                # Escribir los totales en el nuevo archivo Excel
                ws_nuevo["A1"] = "Importe Total"
                ws_nuevo["B1"] = total_importe_total
                ws_nuevo["A2"] = "Animal"
                ws_nuevo["B2"] = "Producto"
                ws_nuevo["C2"] = "Cantidad Vendida"
                ws_nuevo["D2"] = "Precio"
                ws_nuevo["E2"] = "Importe"


                #Aqui tambien va el codigo para escribir los importes individuales por producto
                
                # Escribir los totales en el nuevo archivo Excel
                fila = 3
                for clave, valores in totales.items():
                    ws_nuevo[f"A{fila}"] = clave[0]
                    ws_nuevo[f"B{fila}"] = clave[1]
                    ws_nuevo[f"C{fila}"] = valores[0]  # Cantidad vendida
                    ws_nuevo[f"D{fila}"] = clave[2]
                    ws_nuevo[f'E{fila}']= valores[1]
                    fila += 1

                # Guardar el nuevo archivo Excel
                wb_nuevo.save(ruta_archivo_excel)

                messagebox.showinfo("Éxito", f"Cuadre semanal guardado como {nombre_archivo_excel} en la carpeta 'cuadre semanal'.")

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar el cuadre semanal: {str(e)}")
    
    def cargar_registro_seleccionado(self, treeview, opcion):
        # Función para cargar el/los registro(s) seleccionado(s) en un Toplevel
        if opcion == 'Borrar':
            seleccion = treeview.selection()
            if seleccion:
                # Obtener los valores de los registros seleccionados
                registros_seleccionados = []
                for item in seleccion:
                    valores_fila = treeview.item(item, 'values')
                    registros_seleccionados.append(valores_fila)

                # Mostrar los valores en un Toplevel o procesar la lista según tus necesidades
                return registros_seleccionados
            else:
                # Mostrar un cuadro de diálogo indicando que no se ha seleccionado ningún registro
                messagebox.showinfo("Advertencia", "Seleccione al menos un registro")
                return False
        
        elif opcion=='Actualizar':
            try:
                seleccion=treeview.focus()
            except AttributeError:
                messagebox.showerror('Error','La tabla está vacía')
                raise Exception 
            
            if seleccion:
                # Obtener los valores de la fila seleccionada
                valores_fila = treeview.item(seleccion, 'values')
            
                # Mostrar los valores en un Toplevel
                return valores_fila

            else:
                # Mostrar un cuadro de diálogo indicando que no se ha seleccionado ningún registro
                messagebox.showinfo("Advertencia", "Seleccione un registro")
                return False

    def load_options(self):
        # Obtención de las opciones para el menú desplegable
        with sql.connect(self.archivo) as self.conn:
            self.cursor = self.conn.cursor()
            self.cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table'")
            self.opciones = [fila[0] for fila in self.cursor.fetchall() if fila[0] != 'sqlite_sequence']

    def tablas(self, frame_mostrar, archivo, tabla):
        #Funcion para crear el treeview donde estaran los registros
        def crear_treeview(frame, atributos, valores):
            tree = ttk.Treeview(frame, columns=atributos, show='headings', height=min(len(valores), 10))

            style = ttk.Style()
            style.configure("Treeview", font=('Arial', 14), rowheight=45)
            style.configure("Treeview.Heading", font=('Arial', 14, 'bold'))
            style.configure("Treeview.Treeview", background="#E1E1E1", fieldbackground="#E1E1E1", foreground="black")


            for atributo in atributos:
                tree.heading(atributo, text=atributo)

                # Determina el ancho que tendra una columna en el treeview para que quepa toda la informacion
                ancho = max(tree.heading(atributo)["text"].__len__(), *[len(str(valor[atributos.index(atributo)])) for valor in valores])
                tree.column(atributo, width=ancho * 10)

            for valor in valores:
                tree.insert('', 'end', values=valor)

            return tree

        for widget in frame_mostrar.winfo_children():
            widget.destroy()

        # Muestra un label con la tabla que esta abierta
        titulo_tabla = Label(text=f'Tabla: {tabla}', anchor='w')
        titulo_tabla.config(bg='#082d44', font=(['Arial', 20]), fg='white')
        titulo_tabla.place(
            height=40,
            width=500,
            x=10,
            y=80
        )

        if tabla:
            try:
                with sql.connect(archivo) as conn:
                    cursor = conn.cursor()

                    cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name=?", (tabla,))
                    tabla_existe = cursor.fetchone()

                    if tabla_existe:                        
                        cursor.execute(f"PRAGMA table_info({tabla})")
                        atributos = [column[1] for column in cursor.fetchall()]

                        cursor.execute(f"SELECT * FROM {tabla}")
                        valores = cursor.fetchall()

                        # Crear o actualizar el TreeView
                        if hasattr(self, 'treeview') and not self.treeview.winfo_exists():
                            # Si el TreeView existe pero ya ha sido destruido, crea uno nuevo
                            del self.treeview

                        if not hasattr(self, 'treeview'):
                            # Si no hay un TreeView existente, crea uno nuevo
                            self.treeview = crear_treeview(frame_mostrar, atributos, valores)

                            scrollbar_y = Scrollbar(frame_mostrar, orient="vertical", command=self.treeview.yview)
                            scrollbar_y.pack(side="right", fill="y")
                            self.treeview.configure(yscrollcommand=scrollbar_y.set)

                            scrollbar_x = Scrollbar(frame_mostrar, orient=HORIZONTAL, command=self.treeview.xview)
                            scrollbar_x.pack(side="bottom", fill="x")
                            self.treeview.configure(xscrollcommand=scrollbar_x.set)

                            self.treeview.pack(fill="both", expand=True)

                        return self.treeview
            except TypeError :
                messagebox.showinfo("Alerta", "La tabla está vacía")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo cargar la tabla, error: {str(e)}")
                traceback.print_exc()

    def borrar(self, registros, tabla_actual):
        #Funcion para eliminar el/los registros seleccionados
        while True:
            self.respuesta=respuesta = messagebox.askquestion("Confirmar eliminación", "¿Estás seguro que deseas eliminar?")

            if self.respuesta =='yes':
                break
            elif self.respuesta=='no':
                return
            else:
                continue

        try:
            with sql.connect(self.archivo) as conn:
                cursor = conn.cursor()

                cursor.execute(f"PRAGMA table_info({tabla_actual})")
                info_columnas = cursor.fetchall()
                primary_key_column = next((column[1] for column in info_columnas if column[5] == 1), None)

                if primary_key_column:
                    # Crear la sentencia SQL para borrar el registro
                    sql_query = f"DELETE FROM {tabla_actual} WHERE {primary_key_column} = ?"

                    for registro in registros:
                        # Ejecutar la sentencia SQL
                        cursor.execute(sql_query, (registro[0],))  # Suponiendo que el ID es el primer valor en el registro

                    conn.commit()

                    # Actualizar el TreeView para reflejar los cambios
                    self.actualizar_treeview(tabla_actual)
                    messagebox.showinfo("Éxito", "Registro eliminado correctamente.")
                else:
                    messagebox.showerror("Error", "No se pudo encontrar la clave primaria de la tabla.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo borrar el registro: {str(e)}")

    #Funcion que actualiza el treeview despues de ciertas acciones
    def actualizar_treeview(self, tabla_actual):
        try:
            with sql.connect(self.archivo) as conn:
                cursor = conn.cursor()

                cursor.execute(f"PRAGMA table_info({tabla_actual})")
                atributos = [column[1] for column in cursor.fetchall()]

                cursor.execute(f"SELECT * FROM {tabla_actual}")
                valores = cursor.fetchall()

                # Limpiar el TreeView
                self.treeview.delete(*self.treeview.get_children())

                # Insertar los nuevos datos
                for valor in valores:
                    self.treeview.insert('', 'end', values=valor)
                    
                # Actualizar el monto total si es una tabla de inventario
                if tabla_actual in ["inventario", "inventario_animales"]:
                    total_label = Label(text=f'Monto Total: {self.calcular_monto_total(tabla_actual)}')
                    total_label.config(bg='#082d44', font=(['Arial', 14]), fg='white')
                    total_label.place(
                        height=40,
                        width=500,
                        x=520,
                        y=80
                    )
        except Exception:
            traceback.format_exc()

    def calcular_monto_total(self, tabla_actual):
        if tabla_actual == "inventario" or tabla_actual == "inventario_animales":
            try:
                with sql.connect(self.archivo) as conn:
                    cursor = conn.cursor()

                    cursor.execute(f"SELECT SUM(Importe) FROM {tabla_actual}")
                    total = cursor.fetchone()[0]

                    return total if total else 0  # Si no hay registros, devolver 0
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo calcular el monto total: {str(e)}")
                return 0
        else:
            return 0
        
    def eliminar_todos_los_datos(self, treeview):
        try:
            with sql.connect(self.archivo) as conn:
                cursor = conn.cursor()

                # Eliminar todos los datos de la tabla
                cursor.execute(f"DELETE FROM {self.tabla_actual}")
                conn.commit()

                # Actualizar el TreeView para reflejar los cambios
                self.actualizar_treeview(self.tabla_actual)

                messagebox.showinfo("Éxito", "Todos los datos han sido eliminados correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron eliminar todos los datos: {str(e)}")

    def agrupar_registros(self, tabla_actual):
        # Función para agrupar registros que venden el mismo producto (en inventario) o animal y raza (en inventario_animales)
        try:
            with sql.connect(self.archivo) as conn:
                cursor = conn.cursor()

                # Determinar qué columnas usar para agrupar registros según la tabla actual
                if tabla_actual == "inventario":
                    columnas_agrupacion = "animal, id_producto"
                    query = f"SELECT {columnas_agrupacion}, SUM(cantidad_vendida),precio, SUM(Importe) FROM {tabla_actual} GROUP BY {columnas_agrupacion},articulo"
                elif tabla_actual == "inventario_animales":
                    columnas_agrupacion = "animal, raza"
                    query = f"SELECT {columnas_agrupacion}, SUM(cantidad_vendida),precio, SUM(Importe) FROM {tabla_actual} GROUP BY {columnas_agrupacion}"
                # Consulta SQL para agrupar registros y sumar cantidades vendidas e importes
                cursor.execute(query)
                resultados = cursor.fetchall()

                return resultados
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo agrupar los registros: {str(e)}")
            return []